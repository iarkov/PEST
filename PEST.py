import numpy as np
import pandas as pd
from scipy import stats
import openpyxl

def unify(df):
    dfu = df.copy()
    for j in range(1,len(dfu.columns)):
        array = []

        for i in range(1,len(dfu.index)):
            array.append(abs(dfu.iloc[i-1,j] - dfu.iloc[i,j]))
    
        av = sum(array) / len(array)
    
        ii = 1
        if av < 1:
            while av < 1:
                ii = ii * 10
                av = av * 10  
            for i in range(len(dfu.index)):
                dfu.iloc[i,j] = dfu.iloc[i,j]*ii
        elif av > 1:
            av = int(av)
            avstr = str(av)
            for k in range(len(avstr)-1):
                ii = ii * 10
            for i in range(len(df.index)):
                dfu.iloc[i,j] = dfu.iloc[i,j]/ii
    df = dfu
    dfu = None
    return df

class PEST:
    
    def __init__(self, y, x):
        y = np.array(y)
        y = y.reshape(len(y),1)
        self.y = np.matrix(y)
        onearray = []
        for i in x:
            onearray.append(1)
        X = []
        for i in range(len(onearray)):
            e = []
            e.append(onearray[i])
            for j in x[i]:
                e.append(j)
            X.append(e)
        
        X = np.array(X)
        self.x = np.asmatrix(X)
        self.xt = self.x.transpose()
        self.b = np.dot(self.xt,self.x)
        self.xy = np.dot(self.xt, self.y)
        self.binv = np.linalg.inv(self.b)
        self.a = np.dot(self.binv,self.xy)
    
    
dates = pd.date_range('2004', periods = 13, freq = 'A')
columns = ['Sales, bill. UAH','Inflation, %','Employment pattern, %','GDP per person, USD', 'Currency rate, USD/UAH']
values = np.array([[5.508, 112.3, 71.1, 1367.4, 5.05],
          [7.539, 110.3, 70.9, 1828.7, 5.05],
          [9.401, 111.6, 71.2, 2303, 5.05],
          [12.344, 116.6, 71.7, 3038.6, 5.05],
          [16.121, 122.3, 72.3, 3891, 7.7],
          [20.329, 112.3, 71.6, 2545.5, 8],
          [23.480, 109.1, 71.9, 2974, 7.95],
          [27.213, 104.6, 72.6, 3570.8, 7.99],
          [31.812, 99.8, 72.9, 3856.8, 7.99],
          [35.852, 100.5, 72.9, 4030.3, 7.99],
          [40.8, 124.9, 71.4, 3014.6, 15.7],
          [50, 143.3, 71.5, 2115.4, 26.2],
          [60, 112.4, 71.7, 2354.3, 25.8]])

df = pd.DataFrame(values, index = dates, columns = columns)


##for j in range(1,len(df.columns)):
##    array = []
##
##    for i in range(1,len(df.index)):
##        array.append(abs(df.iloc[i-1,j] - df.iloc[i,j]))
##    
##    av = sum(array) / len(array)
##    
##    ii = 1
##    if av < 1:
##        while av < 1:
##            ii = ii * 10
##            av = av * 10  
##        for i in range(len(df.index)):
##            df.iloc[i,j] = df.iloc[i,j]*ii
##    elif av > 1:
##        av = int(av)
##        avstr = str(av)
##        for k in range(len(avstr)-1):
##            ii = ii * 10
##        for i in range(len(df.index)):
##            df.iloc[i,j] = df.iloc[i,j]/ii

df = unify(df)

book = openpyxl.load_workbook('t.xlsx')
sheet = book.get_sheet_by_name('Feuil1')

values = []
for i in range(7,41):
    e = []
    for j in range(2,6):
        e.append(sheet.cell(row=i, column =j).value)
    values.append(e)

index = []
for i in range(7,41):
    index.append(sheet.cell(row=i, column=1).value)

columns = []
for j in range(2, 6):
    columns.append(sheet.cell(row=5, column=j).value)
    
values = np.array(values)

tdf = pd.DataFrame(values, index = index, columns = columns)

y = np.array(df['Sales, bill. UAH'])
X = np.array(df.drop(['Sales, bill. UAH'], 1))


pest = PEST(y,X)

print(pest.a)

